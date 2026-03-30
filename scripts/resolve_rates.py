#!/usr/bin/env python3
"""
T029 — Rate Resolution Script

Default behavior: preview only (no DB writes).
Use --apply to optionally update consultants.bill_rate / consultants.pay_rate.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import pymysql
import requests


Q4 = Decimal("0.0001")
SPREAD_TOLERANCE = Decimal("0.02")


def q4(value: Optional[Decimal]) -> Optional[Decimal]:
    if value is None:
        return None
    return value.quantize(Q4, rounding=ROUND_HALF_UP)


def dec(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        # Avoid float artifacts by stringifying.
        return Decimal(str(value))
    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None
        s = s.replace("$", "").replace(",", "")
        # Percent forms like "50%" are not money; still parseable if needed.
        if s.endswith("%"):
            s = s[:-1].strip()
            if s == "":
                return None
            return Decimal(s) / Decimal("100")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    return None


_SUFFIX_RE = re.compile(r"\s*\((current|prior)\)\s*$", re.IGNORECASE)


def normalize_name(name: str) -> str:
    s = (name or "").strip().lower()
    s = _SUFFIX_RE.sub("", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def fuzzy_lookup(norm: str, index: Dict[str, Any]) -> Optional[str]:
    """Exact match first; then try prefix match (e.g. 'jagan rao' matches 'jagan rao alleni')."""
    if norm in index:
        return norm
    for key in index:
        if key.startswith(norm + " ") or norm.startswith(key + " "):
            return key
    return None


def parse_env_file(path: Path) -> Dict[str, str]:
    env: Dict[str, str] = {}
    if not path.exists():
        return env
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
            v = v[1:-1]
        env[k] = v
    return env


def find_latest_workbook(search_dir: Path, keyword: str) -> Optional[Path]:
    if not search_dir.exists():
        return None
    keyword_lc = keyword.lower()
    candidates: List[Path] = []
    for p in search_dir.rglob("*.xlsx"):
        name_lc = p.name.lower()
        if keyword_lc in name_lc and "tracking" in name_lc:
            candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


@dataclass(frozen=True)
class RateRow:
    workbook_source: str
    consultant_name: str
    normalized_name: str
    section: str  # fully_known | spread_only
    tax_class: Optional[str]
    pay_rate: Optional[Decimal]
    bill_rate: Optional[Decimal]
    spread_per_hour: Optional[Decimal]
    commission_tier: Optional[Decimal]  # 0.5 for 50%


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_section_header(v: Any, header: str) -> bool:
    return _cell_str(v).strip().lower().startswith(header.lower())


def _header_map(row: List[Any]) -> Dict[str, int]:
    # Map canonical names to indexes by fuzzy matching header strings.
    m: Dict[str, int] = {}
    for idx, v in enumerate(row):
        s = _cell_str(v).lower()
        s = re.sub(r"[^a-z0-9% ]+", " ", s).strip()
        if not s:
            continue
        if s in ("name", "consultant", "consultant name", "employee", "employee name", "full name"):
            m["name"] = idx
        elif "tax" in s and "class" in s:
            m["tax_class"] = idx
        elif "pay" in s and "rate" in s:
            m["pay_rate"] = idx
        elif "bill" in s and "rate" in s:
            m["bill_rate"] = idx
        elif s == "spread" or ("spread" in s and "hour" not in s):
            m["spread_per_hour"] = idx
        elif "spread" in s and "hour" in s:
            m["spread_per_hour"] = idx
        elif "commission" in s and ("tier" in s or "%" in s or "percent" in s):
            m["commission_tier"] = idx
    return m


def _iter_rows(ws) -> Iterable[List[Any]]:
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def read_rate_sheet(workbook_path: Path, workbook_source: str) -> Tuple[Dict[str, RateRow], Dict[str, RateRow], List[RateRow]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Rate Sheet" not in wb.sheetnames:
        raise RuntimeError(f'Workbook "{workbook_path}" missing "Rate Sheet" tab')
    ws = wb["Rate Sheet"]

    fully: Dict[str, RateRow] = {}
    spread_only: Dict[str, RateRow] = {}
    all_rows: List[RateRow] = []

    mode: Optional[str] = None
    header: Dict[str, int] = {}

    def upsert(target: Dict[str, RateRow], rr: RateRow) -> None:
        # Keep "current" row when present, else last row wins.
        existing = target.get(rr.normalized_name)
        if existing is None:
            target[rr.normalized_name] = rr
            return
        cur_flag = "(current)" in rr.consultant_name.lower()
        ex_flag = "(current)" in existing.consultant_name.lower()
        if cur_flag and not ex_flag:
            target[rr.normalized_name] = rr
            return
        if cur_flag == ex_flag:
            target[rr.normalized_name] = rr

    for row in _iter_rows(ws):
        first = row[0] if row else None
        if _is_section_header(first, "FULLY KNOWN RATES"):
            mode = "fully_known"
            header = {}
            continue
        if _is_section_header(first, "SPREAD ONLY"):
            mode = "spread_only"
            header = {}
            continue
        # Terminate data parsing when we hit known non-data sections.
        _first_lc = _cell_str(first).lower()
        if any(_first_lc.startswith(t) for t in ("formula legend", "color code", "key note", "important", "section")):
            mode = None
            continue

        if mode is None:
            continue

        # Find header row for each section.
        if not header:
            hm = _header_map(row)
            if "name" in hm and (("spread_per_hour" in hm) or ("commission_tier" in hm) or ("pay_rate" in hm) or ("bill_rate" in hm)):
                header = hm
            continue

        name_idx = header.get("name")
        if name_idx is None or name_idx >= len(row):
            continue
        name_raw = _cell_str(row[name_idx])
        if name_raw == "":
            # End of section once we hit a fully blank name.
            continue
        # Skip subtotal, formula, and section-label rows.
        _name_lc = name_raw.strip().lower()
        if (
            _name_lc.startswith("total")
            or "subtotal" in _name_lc
            or _name_lc.startswith("formula")
            or _name_lc.startswith("w2 ")
            or _name_lc.startswith("c2c")
            or _name_lc.startswith("ot ")
            or _name_lc.startswith("commission tier")
            or _name_lc.startswith("agency")
            or "%" in name_raw and len(name_raw) < 10  # bare tier labels like "  50% Tier"
        ):
            continue

        norm = normalize_name(name_raw)
        if not norm:
            continue

        tax_class = None
        if "tax_class" in header and header["tax_class"] < len(row):
            tax_class = _cell_str(row[header["tax_class"]]) or None

        pay_rate = None
        bill_rate = None
        spread = None
        tier = None

        if "pay_rate" in header and header["pay_rate"] < len(row):
            pay_rate = q4(dec(row[header["pay_rate"]]))
        if "bill_rate" in header and header["bill_rate"] < len(row):
            bill_rate = q4(dec(row[header["bill_rate"]]))
        if "spread_per_hour" in header and header["spread_per_hour"] < len(row):
            spread = q4(dec(row[header["spread_per_hour"]]))
        if "commission_tier" in header and header["commission_tier"] < len(row):
            tier = dec(row[header["commission_tier"]])
            if isinstance(row[header["commission_tier"]], str) and row[header["commission_tier"]].strip().endswith("%"):
                tier = dec(row[header["commission_tier"]])  # already percent->ratio in dec()
            tier = q4(tier) if tier is not None else None

        rr = RateRow(
            workbook_source=workbook_source,
            consultant_name=name_raw.strip(),
            normalized_name=norm,
            section=mode,
            tax_class=tax_class,
            pay_rate=pay_rate,
            bill_rate=bill_rate,
            spread_per_hour=spread,
            commission_tier=tier,
        )
        all_rows.append(rr)

        if mode == "fully_known":
            upsert(fully, rr)
        else:
            upsert(spread_only, rr)

    return fully, spread_only, all_rows


@dataclass
class LedgerRow:
    workbook_source: str
    consultant_name: str
    tax_class: Optional[str]
    pay_rate: Optional[Decimal]
    bill_rate: Optional[Decimal]
    spread_per_hour: Optional[Decimal]
    commission_tier: Optional[Decimal]
    status: str
    source_am: Optional[str]
    db_id: Optional[int]
    db_current_bill_rate: Optional[Decimal]
    db_current_pay_rate: Optional[Decimal]
    spread_verified: Optional[bool]
    notes: str


def db_connect(env: Dict[str, str]):
    if env.get("DB_CONNECTION", "mysql") != "mysql":
        raise RuntimeError("Only mysql DB_CONNECTION is supported by this script.")
    return pymysql.connect(
        host=env.get("DB_HOST", "127.0.0.1"),
        port=int(env.get("DB_PORT", "3306")),
        user=env.get("DB_USERNAME", "root"),
        password=env.get("DB_PASSWORD", ""),
        database=env.get("DB_DATABASE", ""),
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )


def fetch_consultants(conn) -> List[Dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute("SELECT id, full_name, bill_rate, pay_rate FROM consultants ORDER BY full_name")
        rows = cur.fetchall()
    return rows


def build_db_index(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        norm = normalize_name(r.get("full_name") or "")
        if not norm:
            continue
        idx[norm] = r
    return idx


def fmt_money(v: Optional[Decimal]) -> str:
    if v is None:
        return ""
    return f"{q4(v):f}"


def fmt_tier(v: Optional[Decimal]) -> str:
    if v is None:
        return ""
    return f"{q4(v):f}"


def write_csv(path: Path, header: List[str], rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in header})


def _verify_spread(bill: Decimal, pay: Decimal, tax_class: Optional[str], expected_spread: Decimal) -> Tuple[bool, str]:
    """Check spread reconciliation for both W2 (bill - pay*1.12) and C2C (bill - pay)."""
    tax_lc = (tax_class or "").strip().lower()
    if tax_lc == "w2":
        derived = q4(bill - (pay * Decimal("1.12")))
        formula = "W2: bill-(pay*1.12)"
    else:
        derived = q4(bill - pay)
        formula = "C2C/1099: bill-pay"
    delta = (derived - expected_spread).copy_abs()
    if delta <= SPREAD_TOLERANCE:
        return True, ""
    return False, f"Spread mismatch ({formula}): derived={derived:f} expected={expected_spread:f} delta={delta:f}"


def resolve_cross_workbook(
    sibug_spread: RateRow,
    dim_fully: Dict[str, RateRow],
    har_fully: Dict[str, RateRow],
    dim_spread: Dict[str, RateRow],
    har_spread: Dict[str, RateRow],
) -> Tuple[str, Optional[str], Optional[RateRow], Optional[bool], str]:
    """
    Returns: status, source_am, matched_rate_row, spread_verified, notes
    """
    norm = sibug_spread.normalized_name
    target = None
    source = None

    # Check fully-known first, then spread-only (some AMs have pay/bill in their spread section too).
    # Use fuzzy_lookup to handle minor name variants (e.g. "Jagan Rao" vs "Jagan Rao Alleni").
    k = fuzzy_lookup(norm, dim_fully)
    if k:
        target, source = dim_fully[k], "Dimarumba"
    else:
        k = fuzzy_lookup(norm, har_fully)
        if k:
            target, source = har_fully[k], "Harsono"
        else:
            k = fuzzy_lookup(norm, dim_spread)
            if k and dim_spread[k].pay_rate is not None and dim_spread[k].bill_rate is not None:
                target, source = dim_spread[k], "Dimarumba"
            else:
                k = fuzzy_lookup(norm, har_spread)
                if k and har_spread[k].pay_rate is not None and har_spread[k].bill_rate is not None:
                    target, source = har_spread[k], "Harsono"
                elif fuzzy_lookup(norm, dim_spread) or fuzzy_lookup(norm, har_spread):
                    return "unresolved", None, None, None, "AM workbook has spread-only too (no pay/bill)"
                else:
                    return "unresolved", None, None, None, "Not found in AM workbooks"

    if target.pay_rate is None or target.bill_rate is None:
        return "unresolved", source, None, None, "AM workbook missing pay/bill"

    if sibug_spread.spread_per_hour is None:
        return "spread_mismatch", source, target, False, "Missing Sibug spread_per_hour"

    ok, note = _verify_spread(target.bill_rate, target.pay_rate, target.tax_class, sibug_spread.spread_per_hour)
    if ok:
        return "resolved_cross", source, target, True, ""
    return "spread_mismatch", source, target, False, note


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Resolve consultant pay/bill rates from AM rate sheets (T029).")
    ap.add_argument("--sibug", type=str, default="", help="Path to Sibug workbook (.xlsx).")
    ap.add_argument("--dimarumba", type=str, default="", help="Path to Dimarumba workbook (.xlsx).")
    ap.add_argument("--harsono", type=str, default="", help="Path to Harsono workbook (.xlsx).")
    ap.add_argument("--apply", action="store_true", help="Apply DB updates (requires confirmation).")
    ap.add_argument("--db-json", type=str, default="", help="Path to JSON file with consultant rows (skips live DB).")
    args = ap.parse_args(argv)

    repo_root = Path(__file__).resolve().parents[1]
    payroll_dir = repo_root / "web" / "storage" / "app" / "private" / "uploads" / "payroll"
    out_dir = repo_root / "scripts" / "output"

    sibug_path = Path(args.sibug) if args.sibug else (find_latest_workbook(payroll_dir, "sibug") or Path())
    dim_path = Path(args.dimarumba) if args.dimarumba else (find_latest_workbook(payroll_dir, "dimarumba") or Path())
    har_path = Path(args.harsono) if args.harsono else (find_latest_workbook(payroll_dir, "harsono") or Path())

    missing: List[str] = []
    if not sibug_path or not sibug_path.exists():
        missing.append("Sibug")
    if not dim_path or not dim_path.exists():
        missing.append("Dimarumba")
    if not har_path or not har_path.exists():
        missing.append("Harsono")
    if missing:
        print("ERROR: Missing workbook(s): " + ", ".join(missing))
        print("Provide explicit paths via --sibug/--dimarumba/--harsono, or place files under:")
        print(f"  {payroll_dir}")
        return 2

    dim_fully, dim_spread, dim_all = read_rate_sheet(dim_path, "Dimarumba")
    har_fully, har_spread, har_all = read_rate_sheet(har_path, "Harsono")
    sib_fully, sib_spread, sib_all = read_rate_sheet(sibug_path, "Sibug")

    # Build a unified ledger keyed by normalized name with "best" resolved rate where applicable.
    # Ledger must include all consultants from all three workbooks (rows per workbook entry).
    ledger: List[LedgerRow] = []

    # DB lookup — optional; continues with empty index if DB is unavailable.
    env = parse_env_file(repo_root / "web" / ".env")
    db_index: Dict[str, Any] = {}
    if args.db_json:
        import json as _json
        db_rows = _json.loads(Path(args.db_json).read_text(encoding="utf-8"))
        db_index = build_db_index(db_rows)
        print(f"DB loaded from JSON — {len(db_rows)} consultant(s).")
    else:
        try:
            conn = db_connect(env)
            try:
                db_rows = fetch_consultants(conn)
            finally:
                conn.close()
            db_index = build_db_index(db_rows)
            print(f"DB connected — {len(db_rows)} consultant(s) loaded.")
        except Exception as db_err:
            print(f"WARNING: DB unavailable ({db_err}). CSVs will be written without db_id matches.")

    def attach_db(norm: str) -> Tuple[Optional[int], Optional[Decimal], Optional[Decimal]]:
        key = fuzzy_lookup(norm, db_index)
        r = db_index.get(key) if key else None
        if not r:
            return None, None, None
        return (
            int(r["id"]),
            q4(dec(r.get("bill_rate"))),
            q4(dec(r.get("pay_rate"))),
        )

    resolved_own = 0
    resolved_cross = 0
    unresolved = 0
    spread_mismatch = 0

    def add_from_rate_row(rr: RateRow, status: str, source_am: Optional[str], spread_verified: Optional[bool], notes: str) -> None:
        db_id, db_bill, db_pay = attach_db(rr.normalized_name)
        nonlocal resolved_own, resolved_cross, unresolved, spread_mismatch
        if status == "resolved_own":
            resolved_own += 1
        elif status == "resolved_cross":
            resolved_cross += 1
        elif status == "spread_mismatch":
            spread_mismatch += 1
        elif status == "unresolved":
            unresolved += 1
        ledger.append(
            LedgerRow(
                workbook_source=rr.workbook_source,
                consultant_name=rr.consultant_name,
                tax_class=rr.tax_class,
                pay_rate=rr.pay_rate,
                bill_rate=rr.bill_rate,
                spread_per_hour=rr.spread_per_hour,
                commission_tier=rr.commission_tier,
                status=status,
                source_am=source_am,
                db_id=db_id,
                db_current_bill_rate=db_bill,
                db_current_pay_rate=db_pay,
                spread_verified=spread_verified,
                notes=notes,
            )
        )

    # Fully-known: always resolved_own for each workbook.
    for rr in dim_all:
        if rr.section == "fully_known":
            add_from_rate_row(rr, "resolved_own", "Dimarumba", None, "")
        else:
            add_from_rate_row(rr, "unresolved", None, None, "Spread-only row (not cross-resolved)")

    for rr in har_all:
        if rr.section == "fully_known":
            add_from_rate_row(rr, "resolved_own", "Harsono", None, "")
        else:
            add_from_rate_row(rr, "unresolved", None, None, "Spread-only row (not cross-resolved)")

    # Sibug: fully-known rows are tier=0.5 per spec; spread-only rows are cross-resolved.
    for rr in sib_all:
        if rr.section == "fully_known":
            add_from_rate_row(rr, "resolved_own", "Sibug", None, "")
            continue

        # Spread-only: cross-workbook lookup for pay/bill, verify spread.
        status, source_am, match, spread_ok, notes = resolve_cross_workbook(rr, dim_fully, har_fully, dim_spread, har_spread)
        if status in ("resolved_cross", "spread_mismatch") and match is not None:
            # Fill in pay/bill/tax_class from AM workbook, keep Sibug spread + tier.
            merged = RateRow(
                workbook_source=rr.workbook_source,
                consultant_name=rr.consultant_name,
                normalized_name=rr.normalized_name,
                section=rr.section,
                tax_class=match.tax_class,
                pay_rate=match.pay_rate,
                bill_rate=match.bill_rate,
                spread_per_hour=rr.spread_per_hour,
                commission_tier=rr.commission_tier,
            )
            add_from_rate_row(merged, status, source_am, spread_ok, notes)
        else:
            add_from_rate_row(rr, status, source_am, spread_ok, notes)

    # Write rate-resolution-ledger.csv
    ledger_rows: List[Dict[str, Any]] = []
    for r in ledger:
        ledger_rows.append(
            {
                "workbook_source": r.workbook_source,
                "consultant_name": r.consultant_name,
                "tax_class": r.tax_class or "",
                "pay_rate": fmt_money(r.pay_rate),
                "bill_rate": fmt_money(r.bill_rate),
                "spread_per_hour": fmt_money(r.spread_per_hour),
                "commission_tier": fmt_tier(r.commission_tier),
                "status": r.status,
                "source_am": r.source_am or "",
                "db_id": r.db_id or "",
                "db_current_bill_rate": fmt_money(r.db_current_bill_rate),
                "db_current_pay_rate": fmt_money(r.db_current_pay_rate),
                "spread_verified": ("" if r.spread_verified is None else ("true" if r.spread_verified else "false")),
                "notes": r.notes,
            }
        )

    ledger_rows.sort(key=lambda x: (x["consultant_name"].lower(), x["workbook_source"]))
    write_csv(
        out_dir / "rate-resolution-ledger.csv",
        [
            "workbook_source",
            "consultant_name",
            "tax_class",
            "pay_rate",
            "bill_rate",
            "spread_per_hour",
            "commission_tier",
            "status",
            "source_am",
            "db_id",
            "db_current_bill_rate",
            "db_current_pay_rate",
            "spread_verified",
            "notes",
        ],
        ledger_rows,
    )

    # Build DB update preview (one row per DB id; pick best available resolved row)
    best_by_db_id: Dict[int, LedgerRow] = {}
    for r in ledger:
        if r.db_id is None:
            continue
        if r.status not in ("resolved_own", "resolved_cross"):
            continue
        if r.bill_rate is None or r.pay_rate is None:
            continue
        existing = best_by_db_id.get(r.db_id)
        # Prefer resolved_own over resolved_cross; otherwise first seen.
        if existing is None:
            best_by_db_id[r.db_id] = r
        elif existing.status != "resolved_own" and r.status == "resolved_own":
            best_by_db_id[r.db_id] = r

    preview_rows: List[Dict[str, Any]] = []
    for db_id, r in best_by_db_id.items():
        cur_bill = q4(r.db_current_bill_rate)
        cur_pay = q4(r.db_current_pay_rate)
        prop_bill = q4(r.bill_rate)
        prop_pay = q4(r.pay_rate)
        if cur_bill != prop_bill or cur_pay != prop_pay:
            preview_rows.append(
                {
                    "db_id": db_id,
                    "consultant_name": r.consultant_name,
                    "current_bill_rate": fmt_money(cur_bill),
                    "proposed_bill_rate": fmt_money(prop_bill),
                    "current_pay_rate": fmt_money(cur_pay),
                    "proposed_pay_rate": fmt_money(prop_pay),
                    "source_am": r.source_am or "",
                    "status": r.status,
                }
            )

    preview_rows.sort(key=lambda x: (str(x["consultant_name"]).lower(), int(x["db_id"])))
    write_csv(
        out_dir / "rate-db-update-preview.csv",
        [
            "db_id",
            "consultant_name",
            "current_bill_rate",
            "proposed_bill_rate",
            "current_pay_rate",
            "proposed_pay_rate",
            "source_am",
            "status",
        ],
        preview_rows,
    )

    print(f"Wrote: {out_dir / 'rate-resolution-ledger.csv'}")
    print(f"Wrote: {out_dir / 'rate-db-update-preview.csv'}")
    print("Summary:")
    print(f"  resolved_own: {resolved_own}")
    print(f"  resolved_cross: {resolved_cross}")
    print(f"  unresolved: {unresolved}")
    print(f"  spread_mismatch: {spread_mismatch}")

    if not args.apply:
        return 0

    updates = preview_rows
    print()
    print(f"Preview contains {len(updates)} DB updates.")
    if len(updates) == 0:
        print("Nothing to apply.")
        return 0

    confirm = input(f"Apply {len(updates)} updates to consultants table? (yes/no) ").strip().lower()
    if confirm != "yes":
        print("Aborted.")
        return 1

    conn = db_connect(env)
    log_path = out_dir / "rate-update-log.txt"
    ts = dt.datetime.now().isoformat(timespec="seconds")
    try:
        with conn.cursor() as cur, log_path.open("a", encoding="utf-8") as log:
            log.write(f"\n[{ts}] Applying {len(updates)} consultant rate updates\n")
            for row in updates:
                db_id = int(row["db_id"])
                prop_bill = q4(dec(row["proposed_bill_rate"]))
                prop_pay = q4(dec(row["proposed_pay_rate"]))
                # Re-read current values inside the transaction.
                cur.execute("SELECT bill_rate, pay_rate, full_name FROM consultants WHERE id=%s", (db_id,))
                current = cur.fetchone()
                if not current:
                    log.write(f"- id={db_id}: missing row, skipped\n")
                    continue

                cur_bill = q4(dec(current.get("bill_rate")))
                cur_pay = q4(dec(current.get("pay_rate")))

                set_bill = prop_bill is not None and cur_bill != prop_bill
                set_pay = prop_pay is not None and (cur_pay is None or q4(cur_pay) == Decimal("0.0000"))

                if not set_bill and not set_pay:
                    log.write(f"- id={db_id} ({current.get('full_name')}): no-op (bill/pay already set), skipped\n")
                    continue

                if set_bill and set_pay:
                    cur.execute(
                        "UPDATE consultants SET bill_rate=%s, pay_rate=%s WHERE id=%s",
                        (str(prop_bill), str(prop_pay), db_id),
                    )
                    log.write(
                        f"- id={db_id} ({current.get('full_name')}): bill {cur_bill}->{prop_bill}; pay {cur_pay}->{prop_pay}\n"
                    )
                elif set_bill:
                    cur.execute(
                        "UPDATE consultants SET bill_rate=%s WHERE id=%s",
                        (str(prop_bill), db_id),
                    )
                    log.write(f"- id={db_id} ({current.get('full_name')}): bill {cur_bill}->{prop_bill}\n")
                else:
                    cur.execute(
                        "UPDATE consultants SET pay_rate=%s WHERE id=%s",
                        (str(prop_pay), db_id),
                    )
                    log.write(f"- id={db_id} ({current.get('full_name')}): pay {cur_pay}->{prop_pay}\n")

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # Trigger recompute margins (best-effort; endpoint is admin-only in-app)
    app_url = env.get("APP_URL", "").rstrip("/")
    if app_url:
        try:
            resp = requests.post(f"{app_url}/payroll/recompute-margins", timeout=30)
            with log_path.open("a", encoding="utf-8") as log:
                log.write(f"[{dt.datetime.now().isoformat(timespec='seconds')}] recompute-margins: {resp.status_code}\n")
        except Exception as e:
            with log_path.open("a", encoding="utf-8") as log:
                log.write(f"[{dt.datetime.now().isoformat(timespec='seconds')}] recompute-margins failed: {e}\n")

    print(f"Applied updates. Log: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

