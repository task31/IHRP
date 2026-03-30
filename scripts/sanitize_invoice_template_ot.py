"""
Clear client-specific values from invoice_template_ot.xlsx while keeping
layout, styles, images, and row formulas (OT rate + payroll) intact.

Run from repo root:
  python scripts/sanitize_invoice_template_ot.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = REPO_ROOT / "web" / "storage" / "app" / "templates" / "invoice_template_ot.xlsx"


def main() -> None:
    if not TEMPLATE.is_file():
        raise SystemExit(f"Missing template: {TEMPLATE}")

    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # Header / meta (was filled from sample BridgeBio invoice)
    for coord in ("G4", "G5", "G6"):
        ws[coord].value = None

    # Bill-to block
    for coord in ("C9", "C10", "C11", "C12"):
        ws[coord].value = None

    # Pay period label only — dates filled in per invoice
    ws["E10"].value = "Pay Period:"

    # First detail row (consultant line) — keep F19/G19 formulas
    ws["B19"].value = None
    ws["C19"].value = None
    ws["D19"].value = None
    ws["E19"].value = None

    wb.save(TEMPLATE)
    print(f"Sanitized: {TEMPLATE}")


if __name__ == "__main__":
    main()
