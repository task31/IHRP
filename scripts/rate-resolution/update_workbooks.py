#!/usr/bin/env python3
"""
Phase 8 — Update Rate Sheet tabs from rate-resolution-ledger.csv.

Default: dry-run (no save). Use --apply after confirming the preview; saves workbooks in place.
Never writes formula cells; preserves formatting, dimensions, merges.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

_SUFFIX_RE = re.compile(r"\s*\((current|prior)\)\s*$", re.IGNORECASE)

RESOLVED_STATUSES = frozenset({"resolved_own", "resolved_cross"})


def normalize_name(name: str) -> str:
    s = (name or "").strip().lower()
    s = _SUFFIX_RE.sub("", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_section_header(v: Any, header: str) -> bool:
    return _cell_str(v).strip().lower().startswith(header.lower())


def _header_map_zero_based(row: List[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for idx, v in enumerate(row):
        s = _cell_str(v).lower()
        s = re.sub(r"[^a-z0-9% ]+", " ", s).strip()
        if not s:
            continue
        if s in ("name", "consultant", "consultant name", "employee", "employee name", "full name"):
            m["name"] = idx
        elif "tax" in s and "class" in s:
            m["tax_class"] = idx
        elif "pay" in s and "rate" in s:
            m["pay_rate"] = idx
        elif "bill" in s and "rate" in s:
            m["bill_rate"] = idx
        elif "note" in s or "source" in s:
            m["notes"] = idx
        elif s == "spread" or ("spread" in s and "hour" not in s):
            m["spread_per_hour"] = idx
        elif "spread" in s and "hour" in s:
            m["spread_per_hour"] = idx
        elif "commission" in s and ("tier" in s or "%" in s or "percent" in s):
            m["commission_tier"] = idx
    return m


def _parse_money(s: str) -> Optional[Decimal]:
    if not s or not str(s).strip():
        return None
    t = str(s).strip().replace("$", "").replace(",", "")
    try:
        return Decimal(t)
    except InvalidOperation:
        return None


def load_ledger(path: Path) -> Dict[Tuple[str, str], dict]:
    by_key: Dict[Tuple[str, str], List[dict]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            wb = (row.get("workbook_source") or "").strip()
            name = row.get("consultant_name") or ""
            norm = normalize_name(name)
            if not wb or not norm:
                continue
            by_key[(wb, norm)].append(row)

    best: Dict[Tuple[str, str], dict] = {}
    for k, rows in by_key.items():

        def score(r: dict) -> Tuple[int, int]:
            st = r.get("status") or ""
            pr = _parse_money(r.get("pay_rate") or "")
            br = _parse_money(r.get("bill_rate") or "")
            resolved = 1 if st in RESOLVED_STATUSES and pr is not None and br is not None else 0
            own = 1 if st == "resolved_own" else 0
            return (resolved, own)

        rows_sorted = sorted(rows, key=score, reverse=True)
        best[k] = rows_sorted[0]
    return best


def source_note_resolved(row: dict) -> str:
    st = row.get("status") or ""
    if st == "resolved_own":
        return "Confirmed from Rate Sheet explicit pay/bill pair"
    if st == "resolved_cross":
        am = (row.get("source_am") or "").strip()
        tc = (row.get("tax_class") or "").strip() or "n/a"
        return f"Confirmed via {am} workbook — spread verified ({tc})"
    return ""


SPREAD_ONLY_NOTE = "Spread only — pay/bill not provable from available data"


def iter_rate_sheet_rows(ws) -> Iterable[Tuple[int, List[Any]]]:
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        yield r_i, [c.value for c in row]


def update_one_workbook(
    path: Path,
    workbook_label: str,
    ledger: Dict[Tuple[str, str], dict],
    dry_run: bool,
) -> Tuple[int, List[str]]:
    """
    Returns (cells_updated_or_would_update, messages).
    """
    messages: List[str] = []
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        messages.append(f"ERROR: Could not open {path.name}: {e}")
        return 0, messages

    if "Rate Sheet" not in wb.sheetnames:
        messages.append(f"ERROR: {path.name} missing Rate Sheet tab — skipped")
        return 0, messages

    ws = wb["Rate Sheet"]
    mode: Optional[str] = None
    header: Dict[str, int] = {}
    stats = 0

    for r_i, values in iter_rate_sheet_rows(ws):
        first = values[0] if values else None
        if _is_section_header(first, "FULLY KNOWN RATES"):
            mode = "fully_known"
            header = {}
            continue
        if _is_section_header(first, "SPREAD ONLY"):
            mode = "spread_only"
            header = {}
            continue
        first_lc = _cell_str(first).lower()
        if any(first_lc.startswith(t) for t in ("formula legend", "color code", "key note", "important", "section")):
            mode = None
            continue

        if mode is None:
            continue

        if not header:
            hm = _header_map_zero_based(values)
            if "name" in hm and (
                ("spread_per_hour" in hm) or ("commission_tier" in hm) or ("pay_rate" in hm) or ("bill_rate" in hm)
            ):
                header = hm
            continue

        name_idx = header.get("name")
        if name_idx is None or name_idx >= len(values):
            continue
        name_raw = _cell_str(values[name_idx])
        if name_raw == "":
            continue
        _name_lc = name_raw.strip().lower()
        if (
            _name_lc.startswith("total")
            or "subtotal" in _name_lc
            or _name_lc.startswith("formula")
            or _name_lc.startswith("w2 ")
            or _name_lc.startswith("c2c")
            or _name_lc.startswith("ot ")
            or _name_lc.startswith("commission tier")
            or _name_lc.startswith("agency")
            or ("%" in name_raw and len(name_raw) < 10)
        ):
            continue

        norm = normalize_name(name_raw)
        if not norm:
            continue

        lrow = ledger.get((workbook_label, norm))
        if lrow is None:
            continue

        status = lrow.get("status") or ""
        pay_d = _parse_money(lrow.get("pay_rate") or "")
        bill_d = _parse_money(lrow.get("bill_rate") or "")

        def col_1b(key: str) -> Optional[int]:
            if key not in header:
                return None
            return header[key] + 1

        pay_col = col_1b("pay_rate")
        bill_col = col_1b("bill_rate")
        notes_col = col_1b("notes")

        def safe_write(col: Optional[int], value: Any, label: str) -> None:
            nonlocal stats
            if col is None:
                return
            cell = ws.cell(row=r_i, column=col)
            if cell.data_type == "f":
                messages.append(
                    f"  skip {label} r{r_i}c{col}: formula preserved ({workbook_label} / {name_raw})"
                )
                return
            if dry_run:
                messages.append(f"  would set {label} r{r_i}c{col} = {value} ({name_raw})")
                stats += 1
            else:
                cell.value = value
                stats += 1

        if status in RESOLVED_STATUSES and pay_d is not None and bill_d is not None:
            # Store as float for Excel compatibility
            safe_write(pay_col, float(round(pay_d, 4)), "pay_rate")
            safe_write(bill_col, float(round(bill_d, 4)), "bill_rate")
            note = source_note_resolved(lrow)
            if note and notes_col:
                safe_write(notes_col, note, "notes")
        elif status in ("spread_only", "unresolved"):
            if notes_col:
                safe_write(notes_col, SPREAD_ONLY_NOTE, "notes")

    if not dry_run:
        try:
            wb.save(path)
        except Exception as e:
            messages.append(f"ERROR: Save failed for {path.name}: {e}")
            return 0, messages

    messages.insert(0, f"{stats} cell(s) {'would be updated' if dry_run else 'updated'} in {path.name}")
    return stats, messages


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Fill Rate Sheet pay/bill from rate-resolution-ledger.csv")
    ap.add_argument("--ledger", type=str, required=True, help="Path to rate-resolution-ledger.csv")
    ap.add_argument("--sibug", type=str, required=True, help="Path to Sibug workbook")
    ap.add_argument("--dimarumba", type=str, required=True, help="Path to Dimarumba workbook")
    ap.add_argument("--harsono", type=str, required=True, help="Path to Harsono workbook")
    ap.add_argument("--apply", action="store_true", help="Save workbooks (prompts for confirmation)")
    args = ap.parse_args(argv)

    ledger_path = Path(args.ledger)
    if not ledger_path.is_file():
        print(f"ERROR: Ledger not found: {ledger_path}", file=sys.stderr)
        return 2

    dry_run = not args.apply
    ledger = load_ledger(ledger_path)

    books = [
        ("Sibug", Path(args.sibug)),
        ("Dimarumba", Path(args.dimarumba)),
        ("Harsono", Path(args.harsono)),
    ]

    if not dry_run:
        confirm = input("Apply changes to workbooks on disk? Type yes to save: ").strip().lower()
        if confirm != "yes":
            print("Aborted.")
            return 1

    total = 0
    for label, p in books:
        if not p.exists():
            print(f"ERROR: Missing workbook for {label}: {p}", file=sys.stderr)
            return 2
        n, msgs = update_one_workbook(p, label, ledger, dry_run=dry_run)
        total += n
        for m in msgs:
            print(m)

    mode = "dry-run" if dry_run else "apply"
    print(f"Done ({mode}): {total} total cell touch(es).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
